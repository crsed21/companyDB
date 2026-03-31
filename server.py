#!/usr/bin/env python3
"""
KZ Business Database - Google Drive Server
- Читает все .xlsx из папки Google Drive при старте
- Кеширует данные на 28 секунд, автоматически обновляет по /api/reload
- Сохраняет ВСЕ изменения обратно в database.xlsx на Google Drive
- Эндпоинты: GET /api/load-local, GET /api/reload, POST /api/save-all, POST /api/search
"""

import os, io, json, time
from http.server import HTTPServer, SimpleHTTPRequestHandler
import google.generativeai as genai

# ===================== КЕШ =====================
_cache_data      = None   # список компаний
_cache_timestamp = 0.0    # время последней загрузки
CACHE_TTL        = 28     # секунд — срок жизни кеша (чуть меньше интервала опроса 30 с)

def get_cached():
    """Возвращает (данные, обновлено?). Перечитывает Drive только по истечении TTL."""
    global _cache_data, _cache_timestamp
    age = time.time() - _cache_timestamp
    if _cache_data is not None and age < CACHE_TTL:
        return _cache_data, False
    rows = load_xlsx_from_drive()
    _cache_data      = rows
    _cache_timestamp = time.time()
    return rows, True

def invalidate_cache():
    """Сбрасывает кеш — следующий запрос перечитает Drive."""
    global _cache_data, _cache_timestamp
    _cache_data      = None
    _cache_timestamp = 0.0

# ===================== КОНФИГУРАЦИЯ =====================
API_KEY             = os.environ.get("GEMINI_API_KEY", "")
GDRIVE_FOLDER_ID    = "13SfBTZqqFogzm4j46Dkqp_XUWOTKVerP"   # ID папки из ссылки
SERVICE_ACCOUNT_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "service_account.json"
)
DB_FILENAME = "database.xlsx"   # имя файла для сохранения на Drive

COLUMNS = [
    'Company name', 'Category', 'Status', 'City', 'Website', 'Email',
    'Phone', 'Address', 'CEO-1', 'Position-1', 'CEO-2', 'Position-2',
    'Linkedin', 'Status-L', 'Facebook', 'Status-F'
]

gemini_model = None
drive_service = None


# ===================== GOOGLE DRIVE — СЕРВИС =====================
def get_drive_service():
    """Возвращает авторизованный клиент Google Drive API."""
    global drive_service
    if drive_service:
        return drive_service

    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
    except ImportError:
        raise RuntimeError("pip install google-api-python-client google-auth")

    if not os.path.exists(SERVICE_ACCOUNT_FILE):
        raise FileNotFoundError(
            f"Файл сервисного аккаунта не найден: {SERVICE_ACCOUNT_FILE}\n"
            "Скачай JSON-ключ из Google Cloud Console и положи рядом со скриптом."
        )

    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE,
        scopes=["https://www.googleapis.com/auth/drive"]
    )
    drive_service = build("drive", "v3", credentials=creds, cache_discovery=False)
    print("  ✓ Google Drive API подключён")
    return drive_service


# ===================== GOOGLE DRIVE — ЧТЕНИЕ =====================
def load_xlsx_from_drive():
    """Скачивает все .xlsx из папки Drive, объединяет, возвращает список dict."""
    try:
        import openpyxl
        from googleapiclient.http import MediaIoBaseDownload
    except ImportError:
        print("  ⚠️  pip install openpyxl google-api-python-client")
        return []

    svc = get_drive_service()

    # Ищем все xlsx в папке
    query = (
        f"'{GDRIVE_FOLDER_ID}' in parents "
        "and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' "
        "and trashed=false"
    )
    results = svc.files().list(
        q=query, fields="files(id, name)", pageSize=50
    ).execute()
    files = results.get("files", [])

    if not files:
        print(f"  ⚠️  .xlsx файлы не найдены в папке Drive (ID: {GDRIVE_FOLDER_ID})")
        return []

    COL_MAP = {
        'Company name': ['Company name', 'Название компании', 'название компании'],
        'Category':     ['Category', 'Category ', 'Категория'],
        'Status':       ['Status', 'Status ', 'Статус'],
        'Website':      ['Website', 'Сайт'],
        'Email':        ['Email'],
        'City':         ['City', 'Город'],
        'Phone':        ['Phone', 'Phone, contacts', 'Phone, contacts ', 'Телефон'],
        'Address':      ['Address', 'Адрес'],
        'CEO-1':        ['CEO-1', 'Председатель правления / CEO'],
        'Position-1':   ['Position-1', 'Должность.1'],
        'CEO-2':        ['CEO-2', 'Глава совета директоров'],
        'Position-2':   ['Position-2', 'Должность'],
        'Linkedin':     ['Linkedin', 'LinkedIn'],
        'Status-L':     ['Status-L', 'Статус LinkedIn'],
        'Facebook':     ['Facebook'],
        'Status-F':     ['Status-F', 'Статус Facebook'],
    }

    all_records = []
    seen_names  = set()

    for file_meta in sorted(files, key=lambda f: f["name"]):
        fname   = file_meta["name"]
        file_id = file_meta["id"]
        try:
            # Скачиваем файл в память
            request  = svc.files().get_media(fileId=file_id)
            buf      = io.BytesIO()
            downloader = MediaIoBaseDownload(buf, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()
            buf.seek(0)

            wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
            ws = wb.active
            headers = [
                str(c.value).strip() if c.value else ""
                for c in next(ws.iter_rows(min_row=1, max_row=1))
            ]

            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                raw = {
                    headers[i]: (str(v).strip() if v is not None else "")
                    for i, v in enumerate(row) if i < len(headers)
                }
                rec = {}
                for target, aliases in COL_MAP.items():
                    val = ""
                    for alias in aliases:
                        if raw.get(alias):
                            val = raw[alias]; break
                    rec[target] = val

                name = rec.get('Company name', '').strip()
                if name and name.lower() not in seen_names:
                    seen_names.add(name.lower())
                    all_records.append(rec)
                    count += 1

            wb.close()
            print(f"  ✓ {fname}: {count} записей")
        except Exception as e:
            print(f"  ⚠️  Ошибка {fname}: {e}")

    print(f"  ✓ Итого: {len(all_records)} компаний")
    return all_records


# ===================== GOOGLE DRIVE — ЗАПИСЬ =====================
def save_all_to_drive(records: list):
    """Перезаписывает database.xlsx в папке Drive."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from googleapiclient.http import MediaIoBaseUpload
    except ImportError:
        print("  ⚠️  pip install openpyxl google-api-python-client")
        return False

    svc = get_drive_service()

    # Формируем xlsx в памяти
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Database"

    header_fill = PatternFill("solid", fgColor="C8622A")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for rec in records:
        ws.append([str(rec.get(col, "")) for col in COLUMNS])

    col_widths = {
        'Company name': 35, 'Category': 25, 'Status': 12, 'City': 15,
        'Website': 30, 'Email': 28, 'Phone': 18, 'Address': 35,
        'CEO-1': 25, 'Position-1': 30, 'CEO-2': 25, 'Position-2': 30,
        'Linkedin': 22, 'Status-L': 14, 'Facebook': 22, 'Status-F': 14,
    }
    for i, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = col_widths.get(col, 18)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # Проверяем: существует ли уже database.xlsx в папке?
    query = (
        f"'{GDRIVE_FOLDER_ID}' in parents "
        f"and name='{DB_FILENAME}' "
        "and trashed=false"
    )
    existing = svc.files().list(q=query, fields="files(id)").execute().get("files", [])

    media = MediaIoBaseUpload(buf, mimetype=mime, resumable=False)

    if existing:
        # Обновляем существующий файл
        file_id = existing[0]["id"]
        svc.files().update(
            fileId=file_id,
            media_body=media
        ).execute()
        print(f"  ✓ Обновлено {len(records)} записей → Drive/{DB_FILENAME} (id={file_id})")
    else:
        # Создаём новый файл
        meta = {"name": DB_FILENAME, "parents": [GDRIVE_FOLDER_ID]}
        svc.files().create(
            body=meta, media_body=media, fields="id"
        ).execute()
        print(f"  ✓ Создан {DB_FILENAME} на Drive с {len(records)} записями")

    return True


# ===================== GEMINI =====================
def get_gemini():
    global gemini_model
    if not gemini_model:
        if not API_KEY:
            raise ValueError("GEMINI_API_KEY не задан")
        genai.configure(api_key=API_KEY)
        gemini_model = genai.GenerativeModel(
            model_name="gemini-2.5-flash",
            generation_config={"temperature": 0.1, "response_mime_type": "application/json"}
        )
    return gemini_model


# ===================== HTTP HANDLER =====================
class Handler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(
            *args,
            directory=os.path.dirname(os.path.abspath(__file__)),
            **kwargs
        )

    def log_message(self, format, *args):
        print(f"  {self.command} {self.path.split('?')[0]} → {args[1] if len(args) > 1 else ''}")

    def do_OPTIONS(self):
        self.send_response(200); self._cors(); self.end_headers()

    def do_GET(self):
        if self.path == "/api/load-local":
            self._handle_load()
        elif self.path == "/api/reload":
            self._handle_reload()
        else:
            super().do_GET()

    def do_POST(self):
        routes = {
            "/api/save-all": self._handle_save_all,
            "/api/search":   self._handle_search,
        }
        handler = routes.get(self.path)
        if handler:
            handler()
        else:
            self.send_error(404)

    def _cors(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS, GET")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    def _json_response(self, data, status=200):
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status); self._cors()
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers(); self.wfile.write(body)

    def _read_body(self):
        length = int(self.headers.get("Content-Length", 0))
        return json.loads(self.rfile.read(length))

    # GET /api/load-local — первичная загрузка (всегда свежие данные)
    def _handle_load(self):
        try:
            invalidate_cache()
            rows, _ = get_cached()
            self._json_response({"success": True, "data": rows, "count": len(rows)})
        except Exception as e:
            print(f"  ⚠️  load: {e}")
            self._json_response({"error": str(e)}, 500)

    # GET /api/reload — тихое фоновое обновление (отдаёт из кеша если ещё свежий)
    def _handle_reload(self):
        try:
            rows, refreshed = get_cached()
            if refreshed:
                print(f"  ↻ reload: перечитано {len(rows)} записей")
            self._json_response({"success": True, "data": rows, "count": len(rows), "refreshed": refreshed})
        except Exception as e:
            print(f"  ⚠️  reload: {e}")
            self._json_response({"error": str(e)}, 500)

    # POST /api/save-all
    def _handle_save_all(self):
        try:
            body    = self._read_body()
            records = body.get("data", [])
            ok      = save_all_to_drive(records)
            if ok:
                # Обновляем кеш сохранёнными данными — не нужно перечитывать Drive
                global _cache_data, _cache_timestamp
                _cache_data      = records
                _cache_timestamp = time.time()
            self._json_response({"success": ok, "saved": len(records)})
        except Exception as e:
            print(f"  ⚠️  save-all: {e}")
            self._json_response({"error": str(e)}, 500)

    # POST /api/search
    def _handle_search(self):
        try:
            body         = self._read_body()
            company_name = body.get("company", "").strip()
            categories   = body.get("categories", "")
            if not company_name:
                self._json_response({"error": "Введите название"}, 400); return

            m = get_gemini()
            prompt = f"""You are a professional business researcher.
Return a JSON object about the Kazakhstan company: "{company_name}".
Pick Category from: {categories}.
Use exact keys: "Company name","Category","City","Website","Email","Phone","Address",
"CEO-1","Position-1","CEO-2","Position-2","Linkedin","Facebook","Status-L","Status-F".
For social media use only handles. If not found use "". Return ONLY JSON."""

            resp   = m.generate_content(prompt)
            parsed = json.loads(resp.text.strip())
            parsed["Status"] = "Активный"
            if not parsed.get("Company name"):
                parsed["Company name"] = company_name

            self._json_response({"success": True, "data": parsed})
        except Exception as e:
            print(f"  ⚠️  search: {e}")
            self._json_response({"error": str(e)}, 500)


# ===================== UTILS =====================
def load_env():
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if os.path.exists(path):
        with open(path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    os.environ[k.strip()] = v.strip().strip('"').strip("'")
        print("  ✓ Загружен .env")


# ===================== MAIN =====================
def main():
    load_env()
    global API_KEY
    API_KEY = os.environ.get("GEMINI_API_KEY", API_KEY)

    print("\n" + "=" * 60)
    print("  🗄  KZ Business Database — Google Drive Server")
    print("=" * 60)

    if not API_KEY:
        print("  ⚠️  GEMINI_API_KEY не задан!")
    if not os.path.exists(SERVICE_ACCOUNT_FILE):
        print(f"  ⚠️  service_account.json не найден рядом со скриптом!")

    print(f"\n  ☁️  Drive папка ID: {GDRIVE_FOLDER_ID}")

    try:
        rows = load_xlsx_from_drive()
        print(f"  ✓ Загружено: {len(rows)} компаний\n")
    except Exception as e:
        print(f"  ⚠️  Не удалось загрузить с Drive: {e}\n")

    import socket
    try:
        local_ip = socket.gethostbyname(socket.gethostname())
    except Exception:
        local_ip = "ВАШ_IP"

    port   = 8000
    server = HTTPServer(("0.0.0.0", port), Handler)

    print(f"  ✓ Локально:        http://localhost:{port}")
    print(f"  ✓ По сети (Wi-Fi): http://{local_ip}:{port}")
    print(f"  ✓ Для интернета:   ngrok http {port}")
    print("\n  Ctrl+C для остановки\n")

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n  Сервер остановлен.")


if __name__ == "__main__":
    main()
